## function processing file sheet and dataframe

import pandas as pd
import openpyxl as opxl
from openpyxl.styles import PatternFill
import function_compare as fc_cp

redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
## read xslx file database CE

def read_xlsx(name):
    xl = pd.ExcelFile(name+'.xlsx')
    data = pd.read_excel(xl, 0, header=None)
    return data

# read workbook
def read_excel(file_name):
    workbook = opxl.load_workbook(file_name)
    return workbook

# read worksheet
def read_sheet(workbook):
    worksheet = workbook['5M+E Tool Parameter P3']
    return worksheet

# color cell in the row in which por != actual
def color_cell(worksheet, range_size):
    for i in range(7, range_size):
        if  fc_cp.compare(worksheet.cell(row =i, column =5), worksheet.cell(row =i, column =7),  worksheet.cell(row =i, column =7)) ==0 :
            worksheet.cell(row=i, column=7).fill = redFill

# saving worksheet to db
def save(workbook):
    workbook.save("laser_ce.xlsx")


